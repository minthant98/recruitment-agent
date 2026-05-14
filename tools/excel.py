"""
excel.py — HR audit Excel writer.

Creates and maintains the HR audit sheet.
One row per candidate reviewed. Columns map directly
to the ScreeningResult and JudgeVerdict Pydantic schemas.
"""
import os
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Column definitions ────────────────────────────────
# Order matters — this is the Excel column order
AUDIT_COLUMNS = [
    # Pipeline metadata
    "run_id",
    "reviewed_at",

    # Candidate info
    "candidate_name",
    "candidate_email",
    "total_experience_years",
    "relevant_experience_years",
    "previous_roles",
    "education",

    # JD info
    "jd_title",
    "department",

    # Screening scores
    "required_skills_score",
    "experience_score",
    "preferred_skills_score",
    "education_domain_score",
    "composite_score",
    "experience_match_rating",

    # Screener output
    "matched_required_skills",
    "missing_required_skills",
    "strengths",
    "weaknesses",
    "screener_recommendation",
    "decision_justification",

    # Judge output
    "judge_overall_agrees",
    "judge_suggested_recommendation",
    "conflict_flag",
    "conflict_reason",
    "judge_confidence",

    # HR decision
    "hr_final_decision",
    "hr_reviewed_by",
    "hr_override_reason",
    "was_overridden",

    # Pipeline path
    "hitl_reviewed",
    "pipeline_status",
]

# ── Colour coding by decision ─────────────────────────
DECISION_COLOURS = {
    "SHORTLIST_INTERVIEW": "C6EFCE",   # green
    "HOLD":                "FFEB9C",   # amber
    "REJECT":              "FFC7CE",   # red
}

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=10)
BODY_FONT     = Font(size=10)
BORDER_SIDE   = Side(style="thin", color="D9D9D9")
CELL_BORDER   = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE
)


def _get_or_create_workbook(file_path: str):
    """Load existing workbook or create a new one with headers."""
    path = Path(file_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "HR Audit"
        _write_headers(ws)

    return wb, ws


def _write_headers(ws):
    """Write styled header row."""
    for col_idx, col_name in enumerate(AUDIT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set column widths
    ws.row_dimensions[1].height = 30
    for col_idx in range(1, len(AUDIT_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    # Freeze header row
    ws.freeze_panes = "A2"


def _list_to_str(value) -> str:
    """Convert list to comma-separated string for Excel."""
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    return str(value) if value else ""


def append_audit_row(file_path: str, row_data: dict) -> int:
    """
    Append one candidate row to the audit Excel sheet.

    Args:
        file_path: Path to the Excel file (created if not exists)
        row_data:  Dict with keys matching AUDIT_COLUMNS

    Returns:
        Row number written
    """
    wb, ws = _get_or_create_workbook(file_path)

    # Find next empty row
    next_row = ws.max_row + 1

    # Determine row colour from final decision
    decision = row_data.get("hr_final_decision", "")
    fill_colour = DECISION_COLOURS.get(decision, "FFFFFF")
    row_fill = PatternFill("solid", fgColor=fill_colour)

    # Write each column
    for col_idx, col_name in enumerate(AUDIT_COLUMNS, start=1):
        value = row_data.get(col_name, "")

        # Convert lists to strings
        if isinstance(value, list):
            value = _list_to_str(value)

        # Convert booleans to readable strings
        if isinstance(value, bool):
            value = "Yes" if value else "No"

        cell = ws.cell(row=next_row, column=col_idx, value=value)
        cell.fill      = row_fill
        cell.font      = BODY_FONT
        cell.border    = CELL_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.row_dimensions[next_row].height = 25
    wb.save(file_path)

    return next_row